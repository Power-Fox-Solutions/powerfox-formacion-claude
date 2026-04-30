"""
Genera el Excel de la Matriz de Evaluación de Impactos Ambientales
Tabla N° 15 – Perforación de 3 022 Pozos de Desarrollo – Lote VII/VI
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Matriz Impactos"

# ── Helpers ──────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def font(bold=False, size=8, color="000000"):
    return Font(bold=bold, size=size, color=color, name="Calibri")

def align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(style="thin", color="AAAAAA"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def style_cell(cell, bold=False, size=8, txt_color="000000",
               bg=None, h_align="center", v_align="center", wrap=True):
    cell.font = font(bold=bold, size=size, color=txt_color)
    cell.alignment = align(h=h_align, v=v_align, wrap=wrap)
    if bg:
        cell.fill = fill(bg)
    cell.border = border()

def merge_style(ws, r1, c1, r2, c2, text="", bold=False, size=8,
                txt_color="000000", bg=None, h_align="center"):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1)
    cell.value = text
    style_cell(cell, bold=bold, size=size, txt_color=txt_color,
               bg=bg, h_align=h_align, v_align="center", wrap=True)

# ── Colores ───────────────────────────────────────────────────────────────────
C = {
    "title":     "1F3864",   # azul oscuro
    "act_hdr":   "2E75B6",   # azul medio
    "crit_hdr":  "BDD7EE",   # azul claro
    "fisico":    "D6E4F0",
    "biologico": "D5E8D4",
    "perceptual":"FFE6CC",
    "socio":     "FFF2CC",
    "cultural":  "E1D5E7",
    "medio_lbl": "2E4057",
    "elem_lbl":  "4472C4",
    "white":     "FFFFFF",
    "light":     "F7F7F7",
    "data_neg":  "FCE4D6",   # impacto negativo (valores ≥ 1)
}

# ── Definiciones ──────────────────────────────────────────────────────────────
ACTIVITIES = [
    "Movilización y desmovilización del personal, equipos,\nmaquinarias y materiales",
    "Desbroce y eliminación de cobertura vegetal",
    "Construcción de accesos a las locaciones",
    "Tráfico humano en las locaciones",
    "Movimiento de tierra en las locaciones",
    "Construcción de las plataformas de perforación",
    "Perforación de 3 022 pozos de desarrollo",
    "Construcción e instalación de facilidades de producción\n(baterías, estaciones de compresión, manifolds y planta\nde inyección de agua de formación)",
    "Instalación de líneas de conducción",
]
ACT_NUMS = list(range(1, 10))

CRITERIA = ["I", "EX", "MO", "PE", "RV", "EF", "PR", "AC", "SI", "MC"]
CRITERIA_FULL = [
    "Intensidad", "Extensión", "Momento", "Persistencia", "Reversibilidad",
    "Efecto", "Periodicidad", "Acumulación", "Sinergia", "Recuperabilidad",
]

# (medio, elemento_letra, elemento_nombre, factor_code, factor_nombre, color_key)
FACTORS = [
    ("MEDIO FÍSICO",          "A", "FISIOGRAFÍA",              "A1", "MORFOLOGÍA",                                          "fisico"),
    ("MEDIO FÍSICO",          "A", "FISIOGRAFÍA",              "A2", "DRENAJE",                                             "fisico"),
    ("MEDIO FÍSICO",          "A", "FISIOGRAFÍA",              "A3", "PROCESOS",                                            "fisico"),
    ("MEDIO FÍSICO",          "B", "MICROCLIMA",               "B1", "TEMPERATURA",                                        "fisico"),
    ("MEDIO FÍSICO",          "C", "SUELOS",                   "C1", "CAPA ORGÁNICA Y SUELO MINERAL",                      "fisico"),
    ("MEDIO FÍSICO",          "C", "SUELOS",                   "C2", "CALIDAD DEL SUELO",                                  "fisico"),
    ("MEDIO FÍSICO",          "C", "SUELOS",                   "C3", "FUNCIÓN ECOLÓGICA (BIOPRODUCCIÓN)",                  "fisico"),
    ("MEDIO FÍSICO",          "D", "AGUA",                     "D1", "CALIDAD DE AGUA SUPERFICIAL Y SEDIMENTOS\n(Quebrada \"Pariñas\" y Humedal \"Punta Balcones\")", "fisico"),
    ("MEDIO FÍSICO",          "D", "AGUA",                     "D2", "AGUAS SUBTERRÁNEAS",                                 "fisico"),
    ("MEDIO FÍSICO",          "E", "AIRE",                     "E1", "CALIDAD DEL AIRE",                                   "fisico"),
    ("MEDIO FÍSICO",          "E", "AIRE",                     "E2", "NIVEL DE RUIDO",                                     "fisico"),
    ("MEDIO BIOLÓGICO",       "F", "FLORA Y FAUNA TERRESTRE",  "F1", "VEGETACIÓN TERRESTRE (FORESTA)",                     "biologico"),
    ("MEDIO BIOLÓGICO",       "F", "FLORA Y FAUNA TERRESTRE",  "F2", "VEGETACIÓN TERRESTRE (ARBUSTOS Y HIERBAS)",          "biologico"),
    ("MEDIO BIOLÓGICO",       "F", "FLORA Y FAUNA TERRESTRE",  "F3", "AVES, MAMÍFEROS Y REPTILES",                         "biologico"),
    ("MEDIO BIOLÓGICO",       "G", "FLORA Y FAUNA ACUÁTICA",   "G1", "PECES, PLANCTON Y BENTOS",                           "biologico"),
    ("MEDIO PERCEPTUAL",      "H", "PAISAJE VALOR ESCÉNICO",   "H1", "RECURSO VISUAL (PAISAJISTA)",                        "perceptual"),
    ("MEDIO SOCIO ECONÓMICO", "I", "ECONÓMICO",                "I1", "COMERCIO",                                           "socio"),
    ("MEDIO SOCIO ECONÓMICO", "J", "TERRITORIO",               "J1", "USO DE LA TIERRA",                                   "socio"),
    ("MEDIO SOCIO ECONÓMICO", "K", "POBLACIÓN",                "K1", "CASERÍOS Y CENTROS POBLADOS",                        "socio"),
    ("MEDIO SOCIO ECONÓMICO", "K", "POBLACIÓN",                "K2", "EMPLEO TEMPORAL LOCAL",                              "socio"),
    ("MEDIO CULTURAL",        "L", "CULTURAL",                 "L1", "SITIOS ARQUEOLÓGICOS",                               "cultural"),
]

# ── Datos ─────────────────────────────────────────────────────────────────────
# data[factor_code] = lista de 9 elementos (uno por actividad)
# Cada elemento: lista [I,EX,MO,PE,RV,EF,PR,AC,SI,MC] o None (celda vacía)
DATA = {
    "A1": [                                       # MORFOLOGÍA
        [2,1,4,2,2,1,1,2,None,None],              # Act 1
        None,                                     # Act 2
        [1,1,4,4,2,2,1,1,4,2],                   # Act 3
        None,                                     # Act 4
        [4,1,2,4,2,1,1,4,4,2],                   # Act 5
        [4,1,2,4,2,1,1,4,4,2],                   # Act 6
        None,                                     # Act 7
        None,                                     # Act 8
        [1,3,4,4,4,2,4,2,3,2],                   # Act 9
    ],
    "A2": [                                       # DRENAJE
        [4,2,1,1,4,1,2,2,1,1],                   # Act 1
        None,                                     # Act 2
        [1,2,1,1,2,1,1,1,2,1],                   # Act 3
        None,                                     # Act 4
        [4,2,2,4,2,2,1,1,4,2],                   # Act 5
        [4,2,2,4,2,2,1,1,4,2],                   # Act 6
        None,                                     # Act 7
        None,                                     # Act 8
        [1,3,4,4,4,2,4,2,3,2],                   # Act 9
    ],
    "A3": [                                       # PROCESOS
        [1,1,4,1,2,1,2,4,2,1],                   # Act 1
        None,                                     # Act 2
        [1,2,1,2,4,2,1,1,2,1],                   # Act 3
        None,                                     # Act 4
        [1,1,4,4,4,4,2,2,4,2],                   # Act 5
        [1,1,4,4,4,4,2,2,4,2],                   # Act 6
        None,                                     # Act 7
        None,                                     # Act 8
        [1,4,4,4,4,2,4,2,4,2],                   # Act 9
    ],
    "B1": [                                       # TEMPERATURA
        None,                                     # Act 1
        None,                                     # Act 2
        [2,2,4,1,2,4,4,2,2,1],                   # Act 3
        None,                                     # Act 4
        [1,1,4,2,2,2,1,1,4,4],                   # Act 5
        [4,2,2,None,None,None,None,None,None,None], # Act 6
        None,                                     # Act 7
        None,                                     # Act 8
        [1,3,4,4,3,2,4,2,3,2],                   # Act 9
    ],
    "C1": [                                       # CAPA ORGÁNICA Y SUELO MINERAL
        [1,1,1,1,4,4,4,1,1,2],                   # Act 1
        None,                                     # Act 2
        [1,1,2,2,1,2,1,1,1,1],                   # Act 3
        None,                                     # Act 4
        [1,1,4,4,4,2,2,4,1,2],                   # Act 5
        [1,2,4,4,4,2,2,4,4,2],                   # Act 6
        None,                                     # Act 7
        None,                                     # Act 8
        [2,3,4,4,4,2,4,2,4,2],                   # Act 9
    ],
    "C2": [                                       # CALIDAD DEL SUELO
        [1,2,4,4,4,2,1,1,2,2],                   # Act 1
        None,                                     # Act 2
        [1,2,1,1,1,1,1,1,1,1],                   # Act 3
        None,                                     # Act 4
        [1,4,4,4,2,2,4,1,4,2],                   # Act 5
        [2,4,4,1,4,2,2,4,4,2],                   # Act 6
        [1,2,4,2,2,2,4,4,4,2],                   # Act 7
        None,                                     # Act 8
        [2,3,4,4,4,2,4,2,3,2],                   # Act 9
    ],
    "C3": [                                       # FUNCIÓN ECOLÓGICA
        [1,1,2,2,1,2,1,4,4,2],                   # Act 1
        None,                                     # Act 2
        [2,1,1,4,2,2,4,4,1,2],                   # Act 3
        None,                                     # Act 4
        [1,4,2,4,1,2,4,4,2,2],                   # Act 5
        [4,2,4,2,4,1,2,4,4,2],                   # Act 6
        None,                                     # Act 7
        None,                                     # Act 8
        [1,4,2,4,2,1,4,2,4,2],                   # Act 9
    ],
    "D1": [                                       # CALIDAD AGUA SUPERFICIAL
        None,                                     # Act 1
        None,                                     # Act 2
        [1,1,2,2,1,1,1,4,2,1],                   # Act 3
        None,                                     # Act 4
        [4,1,1,2,1,4,1,4,2,1],                   # Act 5
        None,                                     # Act 6
        None,                                     # Act 7
        None,                                     # Act 8
        None,                                     # Act 9
    ],
    "D2": [                                       # AGUAS SUBTERRÁNEAS
        None,                                     # Act 1
        None,                                     # Act 2
        None,                                     # Act 3
        None,                                     # Act 4
        None,                                     # Act 5
        None,                                     # Act 6
        [4,2,2,6,1,3,4,1,4,2],                   # Act 7
        None,                                     # Act 8
        None,                                     # Act 9
    ],
    "E1": [                                       # CALIDAD DEL AIRE
        [1,1,4,1,1,4,1,4,2,1],                   # Act 1
        [2,2,4,2,2,4,1,4,2,1],                   # Act 2
        [1,1,4,2,1,1,4,2,1,1],                   # Act 3
        None,                                     # Act 4
        [1,1,4,2,1,1,4,2,1,2],                   # Act 5
        [2,4,1,2,4,1,4,1,2,3],                   # Act 6
        [1,3,6,1,4,1,3,6,1,4],                   # Act 7
        [3,2,1,4,2,4,3,2,1,4],                   # Act 8
        [1,2,4,2,4,3,2,1,None,None],             # Act 9
    ],
    "E2": [                                       # NIVEL DE RUIDO
        [1,1,1,1,4,2,4,2,2,4],                   # Act 1
        [1,4,2,1,1,1,4,2,1,2],                   # Act 2
        [1,2,1,1,4,2,4,2,1,2],                   # Act 3
        [2,1,4,2,2,4,1,2,4,1],                   # Act 4
        [1,1,1,4,2,4,2,1,2,4],                   # Act 5
        [4,4,1,1,4,4,4,1,1,4],                   # Act 6
        [1,4,1,4,4,4,4,1,1,4],                   # Act 7
        [4,4,2,1,4,4,2,2,1,4],                   # Act 8
        [1,1,4,4,1,None,None,None,None,None],    # Act 9
    ],
    "F1": [                                       # VEGETACIÓN TERRESTRE FORESTA
        None,                                     # Act 1
        [4,2,4,4,4,2,1,1,4,2],                   # Act 2
        None,                                     # Act 3
        None,                                     # Act 4
        [2,1,4,2,2,1,1,4,1,2],                   # Act 5
        [2,1,4,2,2,1,1,2,4,4],                   # Act 6
        None,                                     # Act 7
        None,                                     # Act 8
        [2,1,1,4,4,4,2,None,None,None],          # Act 9
    ],
    "F2": [                                       # VEGETACIÓN TERRESTRE ARBUSTOS
        None,                                     # Act 1
        [4,2,4,4,2,4,1,4,4,2],                   # Act 2
        None,                                     # Act 3
        None,                                     # Act 4
        [4,2,2,4,2,2,1,1,4,2],                   # Act 5
        [4,2,2,4,2,2,1,1,4,2],                   # Act 6
        None,                                     # Act 7
        None,                                     # Act 8
        [2,1,1,4,4,4,2,None,None,None],          # Act 9
    ],
    "F3": [                                       # AVES MAMÍFEROS Y REPTILES
        [4,2,1,1,4,8,4,1,2,1],                   # Act 1
        None,                                     # Act 2
        [1,4,2,4,1,1,2,2,4,2],                   # Act 3
        None,                                     # Act 4
        [2,4,2,2,2,1,1,6,1,4],                   # Act 5
        [2,2,1,4,2,4,2,1,1,4],                   # Act 6
        [1,2,4,2,4,1,4,2,2,2],                   # Act 7
        None,                                     # Act 8
        [2,3,4,2,3,4,2,2,None,None],             # Act 9
    ],
    "G1": [                                       # PECES PLANCTON Y BENTOS
        None,                                     # Act 1
        [2,2,2,1,1,2,4,4,2,1],                   # Act 2
        [1,1,2,1,1,4,1,4,2,2],                   # Act 3
        None,                                     # Act 4
        None,                                     # Act 5
        None,                                     # Act 6
        None,                                     # Act 7
        [1,1,1,2,1,1,1,4,2,2],                   # Act 8
        None,                                     # Act 9
    ],
    "H1": [                                       # RECURSO VISUAL PAISAJISTA
        [2,1,2,1,4,2,4,4,4,3],                   # Act 1
        [2,4,4,4,4,4,4,2,2,4],                   # Act 2
        [4,4,4,4,4,4,2,2,4,2],                   # Act 3
        None,                                     # Act 4
        [1,2,4,4,4,2,4,4,4,4],                   # Act 5
        [1,4,2,2,4,4,4,4,4,2],                   # Act 6
        None,                                     # Act 7
        [1,1,1,1,1,1,1,1,1,1],                   # Act 8
        [1,3,4,4,3,4,2,4,2,3],                   # Act 9
    ],
    "I1": [                                       # COMERCIO
        [4,4,1,1,4,8,4,1,1,2],                   # Act 1
        [2,4,1,1,4,2,4,4,4,2],                   # Act 2
        [4,4,1,4,2,4,4,4,1,4],                   # Act 3
        [1,2,1,1,4,2,2,2,None,None],             # Act 4
        [1,4,1,1,4,2,4,2,2,2],                   # Act 5
        [1,4,2,1,4,2,4,2,2,2],                   # Act 6
        [1,1,1,4,1,1,4,2,1,1],                   # Act 7
        [1,1,1,4,2,1,2,1,1,2],                   # Act 8
        [1,1,1,4,2,1,2,1,1,2],                   # Act 9
    ],
    "J1": [                                       # USO DE LA TIERRA
        None,                                     # Act 1
        [1,4,2,2,4,4,2,2,None,None],             # Act 2
        [2,1,2,4,4,2,1,2,4,1],                   # Act 3
        None,                                     # Act 4
        [4,1,4,2,2,2,4,4,1,2],                   # Act 5
        None,                                     # Act 6
        None,                                     # Act 7
        [2,4,4,4,2,2,2,1,4,2],                   # Act 8
        [1,3,1,4,1,3,1,3,1,None],                # Act 9
    ],
    "K1": [                                       # CASERÍOS Y CENTROS POBLADOS
        [2,2,1,1,4,4,4,2,1,2],                   # Act 1
        None,                                     # Act 2
        [2,4,1,1,2,1,1,1,1,2],                   # Act 3
        None,                                     # Act 4
        [2,2,1,1,2,2,2,2,2,2],                   # Act 5
        [2,1,1,1,2,2,2,2,1,4],                   # Act 6
        [1,3,1,4,None,5,None,3,1,4],             # Act 7
        [1,1,1,1,1,1,1,None,None,None],          # Act 8
        [1,1,1,1,1,4,None,None,None,None],       # Act 9
    ],
    "K2": [                                       # EMPLEO TEMPORAL LOCAL
        [2,2,4,1,4,4,4,1,1,2],                   # Act 1
        [2,4,1,1,4,4,1,1,2,2],                   # Act 2
        [2,1,1,1,1,1,1,1,1,3],                   # Act 3
        [2,6,1,1,1,1,2,1,None,None],             # Act 4
        [1,1,4,1,4,1,1,2,2,2],                   # Act 5
        [2,2,2,1,4,1,4,2,2,2],                   # Act 6
        [2,2,1,1,4,1,4,2,2,2],                   # Act 7
        [2,2,4,2,2,2,4,2,2,2],                   # Act 8
        [1,1,1,1,None,None,None,None,None,None], # Act 9
    ],
    "L1": [                                       # SITIOS ARQUEOLÓGICOS
        None,                                     # Act 1
        None,                                     # Act 2
        [1,4,1,2,2,4,4,4,1,4],                   # Act 3
        None,                                     # Act 4
        [4,2,4,1,4,4,4,1,4,2],                   # Act 5
        None,                                     # Act 6
        None,                                     # Act 7
        None,                                     # Act 8
        [4,1,4,4,4,1,4,2,4,None],                # Act 9
    ],
}

# ── Layout ────────────────────────────────────────────────────────────────────
# Columnas fijas: 1=MEDIO, 2=ELEMENTO(letra), 3=ELEMENTO(nombre), 4=FACTOR(código), 5=FACTOR(nombre)
FIXED_COLS = 5
# Columnas de datos: 9 actividades × 10 criterios = 90
TOTAL_DATA_COLS = 9 * 10
TOTAL_COLS = FIXED_COLS + TOTAL_DATA_COLS  # 95

# Filas
ROW_TITLE   = 1
ROW_ACT_NUM = 2   # "1", "2", … "9"
ROW_ACT_NAM = 3   # nombre de la actividad
ROW_CRIT    = 4   # I EX MO PE RV EF PR AC SI MC (× 9)
ROW_LEGEND  = 5   # leyenda de criterios
DATA_START  = 6   # primera fila de datos

# ── Anchuras de columna ───────────────────────────────────────────────────────
ws.column_dimensions[get_column_letter(1)].width = 14  # MEDIO
ws.column_dimensions[get_column_letter(2)].width = 5   # letra elem.
ws.column_dimensions[get_column_letter(3)].width = 16  # nombre elem.
ws.column_dimensions[get_column_letter(4)].width = 5   # código factor
ws.column_dimensions[get_column_letter(5)].width = 28  # nombre factor
for c in range(FIXED_COLS + 1, TOTAL_COLS + 1):
    ws.column_dimensions[get_column_letter(c)].width = 3.2

# ── Alturas de fila ────────────────────────────────────────────────────────────
ws.row_dimensions[ROW_TITLE].height   = 28
ws.row_dimensions[ROW_ACT_NUM].height = 15
ws.row_dimensions[ROW_ACT_NAM].height = 68
ws.row_dimensions[ROW_CRIT].height    = 28
ws.row_dimensions[ROW_LEGEND].height  = 60

# ── FILA 1: Título ────────────────────────────────────────────────────────────
merge_style(ws, ROW_TITLE, 1, ROW_TITLE, TOTAL_COLS,
            text="TABLA N° 15  –  MATRIZ DE EVALUACIÓN DE IMPACTOS AMBIENTALES\n"
                 "PERFORACIÓN DE 3 022 POZOS DE DESARROLLO  –  LOTE VII/VI",
            bold=True, size=10, txt_color="FFFFFF", bg=C["title"])

# ── FILA 2: Encabezados de columnas fijas y numeración de actividades ─────────
for c, label in enumerate(["MEDIO\nAMBIENTE", "EL.", "ELEMENTOS\nAMBIENTALES",
                            "F.", "FACTORES\nAMBIENTALES"], start=1):
    merge_style(ws, ROW_ACT_NUM, c, ROW_CRIT, c,
                text=label, bold=True, size=8, txt_color="FFFFFF", bg=C["act_hdr"])

for i, act_num in enumerate(ACT_NUMS):
    col_start = FIXED_COLS + i * 10 + 1
    col_end   = col_start + 9
    merge_style(ws, ROW_ACT_NUM, col_start, ROW_ACT_NUM, col_end,
                text=str(act_num), bold=True, size=9, txt_color="FFFFFF", bg=C["act_hdr"])

# ── FILA 3: Nombres de actividades ────────────────────────────────────────────
ws.row_dimensions[ROW_ACT_NUM].height = 14
for i, act_name in enumerate(ACTIVITIES):
    col_start = FIXED_COLS + i * 10 + 1
    col_end   = col_start + 9
    merge_style(ws, ROW_ACT_NAM, col_start, ROW_ACT_NAM, col_end,
                text=act_name, bold=False, size=7, txt_color="FFFFFF", bg=C["act_hdr"])

# ── FILA 4: Criterios ─────────────────────────────────────────────────────────
for i in range(9):
    for j, crit in enumerate(CRITERIA):
        c = FIXED_COLS + i * 10 + j + 1
        cell = ws.cell(ROW_CRIT, c)
        cell.value = crit
        style_cell(cell, bold=True, size=7, txt_color="000000", bg=C["crit_hdr"])

# ── FILA 5: Leyenda de criterios ──────────────────────────────────────────────
legend_text = (
    "I = Intensidad  |  EX = Extensión  |  MO = Momento  |  PE = Persistencia  |  "
    "RV = Reversibilidad  |  EF = Efecto  |  PR = Periodicidad  |  "
    "AC = Acumulación  |  SI = Sinergia  |  MC = Recuperabilidad"
)
merge_style(ws, ROW_LEGEND, 1, ROW_LEGEND, TOTAL_COLS,
            text=legend_text, bold=False, size=8, bg=C["light"], h_align="left")

# ── Filas de datos ────────────────────────────────────────────────────────────
# Agrupar factores por medio y elemento para merge
from itertools import groupby

medio_groups  = {}  # medio -> list of row indices
elem_groups   = {}  # (medio, elem_letra) -> list of row indices

for row_i, factor in enumerate(FACTORS):
    medio, elem_l, elem_n, fc, fn, ck = factor
    medio_groups.setdefault(medio, []).append(row_i)
    elem_groups.setdefault((medio, elem_l, elem_n), []).append(row_i)

# Crear filas
for row_i, factor in enumerate(FACTORS):
    medio, elem_l, elem_n, fc, fn, ck = factor
    xls_row = DATA_START + row_i
    ws.row_dimensions[xls_row].height = 22
    bg = C[ck]

    # Col 4: código factor
    cell = ws.cell(xls_row, 4)
    cell.value = fc
    style_cell(cell, bold=True, size=8, bg=bg)

    # Col 5: nombre factor
    cell = ws.cell(xls_row, 5)
    cell.value = fn
    style_cell(cell, bold=False, size=7, bg=bg, h_align="left")

    # Datos
    factor_data = DATA.get(fc, [None] * 9)
    for act_i, act_vals in enumerate(factor_data):
        for crit_i in range(10):
            c = FIXED_COLS + act_i * 10 + crit_i + 1
            cell = ws.cell(xls_row, c)
            if act_vals is not None and crit_i < len(act_vals) and act_vals[crit_i] is not None:
                cell.value = act_vals[crit_i]
                style_cell(cell, size=7, bg="FDEBD0" if act_vals[crit_i] >= 1 else C["white"])
            else:
                style_cell(cell, size=7, bg=C["white"])

# Merge col 1 (MEDIO)
for medio, rows in medio_groups.items():
    r1 = DATA_START + rows[0]
    r2 = DATA_START + rows[-1]
    # MEDIO
    if r1 == r2:
        cell = ws.cell(r1, 1)
        cell.value = medio
        style_cell(cell, bold=True, size=8, txt_color="FFFFFF", bg=C["medio_lbl"])
    else:
        ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
        cell = ws.cell(r1, 1)
        cell.value = medio
        style_cell(cell, bold=True, size=8, txt_color="FFFFFF", bg=C["medio_lbl"])

# Merge col 2 y 3 (ELEMENTO letra y nombre)
for (medio, elem_l, elem_n), rows in elem_groups.items():
    r1 = DATA_START + rows[0]
    r2 = DATA_START + rows[-1]
    ck = FACTORS[rows[0]][5]
    bg = C[ck]
    for c_idx, val in [(2, elem_l), (3, elem_n)]:
        if r1 == r2:
            cell = ws.cell(r1, c_idx)
            cell.value = val
            style_cell(cell, bold=True, size=8, txt_color="FFFFFF", bg=C["elem_lbl"])
        else:
            ws.merge_cells(start_row=r1, start_column=c_idx, end_row=r2, end_column=c_idx)
            cell = ws.cell(r1, c_idx)
            cell.value = val
            style_cell(cell, bold=True, size=8, txt_color="FFFFFF", bg=C["elem_lbl"])

# ── Inmovilizar paneles ───────────────────────────────────────────────────────
ws.freeze_panes = ws.cell(DATA_START, FIXED_COLS + 1)

# ── Hoja 2: Leyenda completa ──────────────────────────────────────────────────
ws2 = wb.create_sheet("Leyenda y Criterios")
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 60
ws2.column_dimensions["C"].width = 20

headers = [("SIGLA", "DESCRIPCIÓN", "ESCALA DE VALORES")]
for col_i, h in enumerate(headers[0], start=1):
    cell = ws2.cell(1, col_i)
    cell.value = h
    style_cell(cell, bold=True, size=10, txt_color="FFFFFF", bg=C["title"])

criteria_desc = [
    ("I – Intensidad",        "Grado de incidencia de la acción sobre el factor. Varía entre 1 (mínima) y 4 (máxima).", "1=Baja  2=Media  3=Alta  4=Muy Alta"),
    ("EX – Extensión",        "Área de influencia del impacto respecto al entorno del proyecto.",                        "1=Puntual  2=Parcial  4=Extenso"),
    ("MO – Momento",          "Plazo de manifestación del impacto.",                                                     "1=Largo  2=Medio  4=Inmediato"),
    ("PE – Persistencia",     "Permanencia del efecto desde la aparición hasta que se retorne a la línea base.",         "1=Fugaz  2=Temporal  4=Permanente"),
    ("RV – Reversibilidad",   "Posibilidad de retornar a la situación anterior por medios naturales.",                   "1=Corto  2=Medio  4=Irreversible"),
    ("EF – Efecto",           "Forma de manifestación del impacto.",                                                     "1=Indirecto  4=Directo"),
    ("PR – Periodicidad",     "Regularidad de la manifestación del efecto.",                                             "1=Irregular  2=Periódico  4=Continuo"),
    ("AC – Acumulación",      "Incremento progresivo de la manifestación del efecto cuando persiste de forma continua.", "1=Simple  4=Acumulativo"),
    ("SI – Sinergia",         "Reforzamiento de dos o más efectos simples.",                                             "1=Sin sinergia  2=Sinérgico  4=Muy sinérgico"),
    ("MC – Recuperabilidad",  "Posibilidad de retornar a la situación anterior mediante intervención humana.",           "1=Recuperable  2=Mitigable  4=Irrecuperable"),
]

for row_i, (sigla, desc, escala) in enumerate(criteria_desc, start=2):
    ws2.row_dimensions[row_i].height = 22
    bg_row = "F7F7F7" if row_i % 2 == 0 else "FFFFFF"
    for col_i, val in enumerate([sigla, desc, escala], start=1):
        cell = ws2.cell(row_i, col_i)
        cell.value = val
        cell.font = Font(size=9, name="Calibri")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = fill(bg_row)
        cell.border = border()

# ── Guardar ───────────────────────────────────────────────────────────────────
out_path = r"c:\trabajos\claudecode\powerfox-formacion-claude\alumnos\luis-monroy\Tabla15_Matriz_Impactos_Ambientales.xlsx"
wb.save(out_path)
print(f"Archivo guardado: {out_path}")
